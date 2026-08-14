r"""Render a survey spec as a workbook, and read one back.

The workbook is what an RA sees. It is not the source of truth - ``spec.json``
is, because that is what git can diff and a reviewer can read - but it is the
*editing* surface, which makes the direction that matters xlsx -> json. An edit
somebody makes in Excel and that does not survive the trip back is silent data
loss on the instrument, so :func:`read_xlsx` and :func:`write_xlsx` are tested
as a round trip rather than separately.

Sheets, matching SurveyCTO's own workbook so the shape is recognised on sight:

``survey``
    One row per question, including the whole subgraph it becomes.
``choices``
    One row per option. The 24-character cap lives here.
``messages``
    The strings with no position in the graph - the retry nudges, the stop words.
``settings``
    One row describing the instrument.
``help-*``
    Instructions, shipped inside the file. SurveyCTO does this and it is the
    reason its format is learnable without a wiki: the guidance travels with the
    thing being filled in.

Excel is a hostile serialisation format, and each of these is a real defect this
module handles rather than a hypothetical:

* **A label like ``1-2 times`` is a date.** Excel coerces it on the way in, and
  in some locales on the way out, so every text column is written with the ``@``
  number format and read with ``dtype=str``.
* **``option_id`` ``p1_0``, ``value`` ``-99``, a caseid ``007``.** Identifiers,
  not numbers. Same fix, and the same lesson ``push_file`` learned the hard way.
* **Newlines.** Question bodies carry ``\\n\\n``. Excel stores them literally,
  but an edit on Windows can hand back ``\\r\\n``, which would change both the
  flow definition and the message a respondent reads. Normalised on read.
* **Trailing whitespace.** Invisible in a cell, and load-bearing in a label that
  is compared literally against a reply. Stripped on read.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .spec import (
    CONSTRAINED_TYPES,
    QUESTION_TYPES,
    ROLES,
    STRUCTURE_TYPES,
    ChoiceRow,
    MessageRow,
    Settings,
    Spec,
    SpecError,
    SurveyRow,
)

SURVEY_SHEET = "survey"
CHOICES_SHEET = "choices"
MESSAGES_SHEET = "messages"
SETTINGS_SHEET = "settings"

#: Columns whose value is one string per language, rendered as ``label:en``.
_TRANSLATABLE = {
    SURVEY_SHEET: ("label", "list_button", "template"),
    CHOICES_SHEET: ("label", "description", "typed"),
    MESSAGES_SHEET: ("text",),
}

#: Plain columns, in the order they appear in each sheet. Order is part of the
#: interface: an RA reads left to right, so what a question *is* comes before how
#: it behaves, and the rarely-touched flags come last.
_SURVEY_COLUMNS = ("type", "name", "role", "relevance")
_SURVEY_TAIL = (
    "retries",
    "timeout",
    "constraint",
    "constraint_message",
    "stop_check",
    "publish",
    "encrypt",
)
_CHOICES_COLUMNS = ("list_name", "value", "option_id")
_MESSAGES_COLUMNS = ("key",)

#: Written by :func:`write_xlsx`, ignored by :func:`read_xlsx`. It is the claim
#: the spec makes - a row is a subgraph, not a line of configuration - made
#: visible next to the row that makes it. Derived, so it is never read back.
_WIDGETS_COLUMN = "widgets"

_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_DERIVED_FILL = PatternFill("solid", fgColor="F2F2F2")
_TEXT_FORMAT = "@"

#: Wide enough to read a question body without opening the formula bar, narrow
#: enough that the behaviour columns stay on screen.
_WIDTHS = {
    "type": 22,
    "name": 16,
    "role": 12,
    "relevance": 22,
    "list_name": 12,
    "value": 8,
    "option_id": 14,
    "key": 20,
    _WIDGETS_COLUMN: 9,
}
_LABEL_WIDTH = 52
_DEFAULT_WIDTH = 14


def _flatten(value: Any) -> str:
    """Render a cell value as the text Excel should hold."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    return str(value)


def _clean_cell(value: Any) -> str:
    r"""Normalise one cell read back out of a workbook.

    ``\\r\\n`` becomes ``\\n`` because an edit on Windows can introduce it and it
    would otherwise change the message a respondent receives. Trailing
    whitespace goes because it is invisible in a cell and load-bearing in a
    label, which is compared literally against the reply.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value)
    if text.lower() in ("nan", "none"):
        return ""
    return text.replace("\r\n", "\n").replace("\r", "\n").strip()


def _parse_bool(value: str, default: bool) -> bool:
    """Read a yes/no cell, tolerating what people actually type."""
    text = _clean_cell(value).casefold()
    if not text:
        return default
    if text in ("yes", "y", "true", "1", "si", "sí"):
        return True
    if text in ("no", "n", "false", "0"):
        return False
    return default


def _parse_int(value: str) -> int | None:
    """Read a whole-number cell, or None if it is blank or not a number.

    Excel hands back ``2.0`` for a cell holding ``2``, so this goes via float.
    """
    text = _clean_cell(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _languages(columns: list[str], stem: str) -> list[str]:
    """Find the language codes present as ``<stem>:<lang>`` columns."""
    prefix = f"{stem}:"
    return [c[len(prefix) :] for c in columns if c.startswith(prefix)]


def _text_from(row: dict[str, Any], stem: str, languages: list[str]) -> dict[str, str]:
    """Collect ``<stem>:<lang>`` cells into a translatable value."""
    collected = {}
    for lang in languages:
        text = _clean_cell(row.get(f"{stem}:{lang}"))
        if text:
            collected[lang] = text
    return collected


# ---------------------------------------------------------------------------
# Writing.
# ---------------------------------------------------------------------------


def _headers(spec: Spec, sheet: str) -> list[str]:
    """Build one sheet's header row: plain columns, then one per language."""
    languages = spec.settings.languages
    if sheet == SURVEY_SHEET:
        columns = list(_SURVEY_COLUMNS)
        for stem in _TRANSLATABLE[SURVEY_SHEET]:
            columns += [f"{stem}:{lang}" for lang in languages]
        return columns + list(_SURVEY_TAIL) + [_WIDGETS_COLUMN]
    if sheet == CHOICES_SHEET:
        columns = list(_CHOICES_COLUMNS)
        for stem in _TRANSLATABLE[CHOICES_SHEET]:
            columns += [f"{stem}:{lang}" for lang in languages]
        return columns
    columns = list(_MESSAGES_COLUMNS)
    for stem in _TRANSLATABLE[MESSAGES_SHEET]:
        columns += [f"{stem}:{lang}" for lang in languages]
    return columns


def _survey_values(spec: Spec, row: SurveyRow, headers: list[str]) -> list[str]:
    """Render one survey row against the header order."""
    plain: dict[str, Any] = {
        "type": row.type,
        "name": row.name,
        "role": row.role,
        "relevance": row.relevance,
        "retries": row.retries,
        "timeout": row.timeout,
        "constraint": row.constraint,
        "constraint_message": row.constraint_message,
        "stop_check": row.stop_check,
        "publish": row.publish,
        "encrypt": row.encrypt,
        _WIDGETS_COLUMN: spec.widget_count(row),
    }
    translatable = {
        "label": row.label,
        "list_button": row.list_button,
        "template": row.template,
    }
    return [_cell_for(header, plain, translatable) for header in headers]


def _cell_for(
    header: str, plain: dict[str, Any], translatable: dict[str, dict[str, str]]
) -> str:
    """Pick the value for one header, whether plain or ``stem:lang``."""
    if ":" in header:
        stem, _, lang = header.partition(":")
        return _flatten(translatable.get(stem, {}).get(lang, ""))
    return _flatten(plain.get(header, ""))


def _write_sheet(
    workbook: Workbook, title: str, headers: list[str], rows: list[list[str]]
) -> None:
    """Write one sheet, formatted so it can be read and edited without surprises."""
    sheet = workbook.create_sheet(title)
    sheet.append(headers)

    for values in rows:
        sheet.append(values)

    for index, header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        head = sheet.cell(row=1, column=index)
        head.font = Font(bold=True)
        head.fill = _DERIVED_FILL if header == _WIDGETS_COLUMN else _HEADER_FILL
        head.alignment = Alignment(vertical="top", wrap_text=True)

        stem = header.partition(":")[0]
        sheet.column_dimensions[letter].width = _WIDTHS.get(
            header,
            _LABEL_WIDTH
            if stem in ("label", "text", "description")
            else _DEFAULT_WIDTH,
        )

        # Every cell is text. This is the fix for `1-2 times` arriving back as a
        # date and for `007` arriving back as 7 - the cap and the join key
        # respectively, both silently wrong and neither visible in the sheet.
        for cell in sheet[letter][1:]:
            cell.number_format = _TEXT_FORMAT
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # The header stays put while somebody scrolls a 200-row instrument.
    sheet.freeze_panes = "A2"


def _add_dropdowns(sheet, headers: list[str], choice_lists: list[str]) -> None:
    """Constrain `type` and `role` to what the compiler understands.

    Typos in these two are the ones with no good failure mode: a misspelled type
    is caught by `check_spec`, but only after somebody has filled in the row and
    wondered why. A dropdown moves that to the moment of typing.
    """
    types = sorted(STRUCTURE_TYPES) + sorted(QUESTION_TYPES - {"numeric"})
    with_lists = [
        f"{kind} {name}"
        for kind in ("select_list", "select_button")
        for name in sorted(choice_lists)
    ]
    options = types + with_lists

    # Excel refuses an inline list longer than 255 characters, and silently drops
    # the validation rather than reporting it - so it is left off instead of
    # shipping a workbook whose dropdown quietly does nothing.
    joined = ",".join(options)
    if len(joined) <= 255:
        validation = DataValidation(
            type="list", formula1=f'"{joined}"', allow_blank=True
        )
        validation.error = "Not a type this compiler understands."
        validation.promptTitle = "Question type"
        validation.prompt = "How the question is asked, and what subgraph it becomes."
        sheet.add_data_validation(validation)
        column = get_column_letter(headers.index("type") + 1)
        validation.add(f"{column}2:{column}500")

    roles = ",".join(sorted(r for r in ROLES if r))
    if "role" in headers and len(roles) <= 255:
        role_validation = DataValidation(
            type="list", formula1=f'"{roles}"', allow_blank=True
        )
        role_validation.promptTitle = "Role"
        role_validation.prompt = (
            "Leave blank for an ordinary question. Only the spine positions - the "
            "opener, consent, a closing - need a role."
        )
        sheet.add_data_validation(role_validation)
        column = get_column_letter(headers.index("role") + 1)
        role_validation.add(f"{column}2:{column}500")


def _write_settings(workbook: Workbook, spec: Spec) -> None:
    """Write the settings sheet as key/value pairs, one row each.

    Key/value rather than SurveyCTO's one-wide-row, because this carries
    per-language values and a row of forty columns cannot be read. The keys are
    SurveyCTO's where SurveyCTO has one.
    """
    settings = spec.settings
    rows: list[tuple[str, str]] = [
        ("form_id", settings.form_id),
        ("form_title", settings.form_title),
        ("languages", ", ".join(settings.languages)),
        ("default_language", settings.default_language),
        ("functions_service", settings.functions_service),
        ("default_timeout", str(settings.default_timeout)),
        ("default_retries", str(settings.default_retries)),
        ("preloads", ", ".join(settings.preloads)),
    ]
    for lang in settings.languages:
        rows.append((f"flow_name:{lang}", settings.flow_name.get(lang, "")))
    for lang in settings.languages:
        rows.append((f"description:{lang}", settings.description.get(lang, "")))

    _write_sheet(workbook, SETTINGS_SHEET, ["key", "value"], [[k, v] for k, v in rows])


def write_xlsx(spec: Spec, path: Path) -> Path:
    """Write a spec as a workbook.

    Args:
        spec: The instrument.
        path: Destination ``.xlsx``. Overwritten if it exists, because this is a
            generated view of the JSON rather than a document with a history.

    Returns:
        The path written.

    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    survey_headers = _headers(spec, SURVEY_SHEET)
    _write_sheet(
        workbook,
        SURVEY_SHEET,
        survey_headers,
        [_survey_values(spec, row, survey_headers) for row in spec.survey],
    )
    _add_dropdowns(
        workbook[SURVEY_SHEET],
        survey_headers,
        sorted({c.list_name for c in spec.choices}),
    )

    choices_headers = _headers(spec, CHOICES_SHEET)
    _write_sheet(
        workbook,
        CHOICES_SHEET,
        choices_headers,
        [
            [
                _cell_for(
                    header,
                    {
                        "list_name": choice.list_name,
                        "value": choice.value,
                        "option_id": choice.option_id,
                    },
                    {
                        "label": choice.label,
                        "description": choice.description,
                        "typed": choice.typed,
                    },
                )
                for header in choices_headers
            ]
            for choice in spec.choices
        ],
    )

    messages_headers = _headers(spec, MESSAGES_SHEET)
    _write_sheet(
        workbook,
        MESSAGES_SHEET,
        messages_headers,
        [
            [
                _cell_for(header, {"key": message.key}, {"text": message.text})
                for header in messages_headers
            ]
            for message in spec.messages
        ],
    )

    _write_settings(workbook, spec)
    _write_help(workbook, spec)

    workbook.save(path)
    return path


# ---------------------------------------------------------------------------
# The help sheets, shipped inside the file.
# ---------------------------------------------------------------------------


def _help_survey(spec: Spec) -> list[str]:
    """Return the guidance for the survey sheet."""
    return [
        "The survey worksheet",
        "",
        "One row is one question - and one row is a whole subgraph in the flow, "
        "not a line of configuration. The `widgets` column on the right says how "
        "many Studio widgets that row becomes. It is greyed out because it is "
        "derived: editing it does nothing.",
        "",
        "type - how the question is asked, and therefore what it becomes",
        "  text                 a plain message; ANY reply is stored as the answer",
        "  integer / decimal    a plain message; the reply must match `constraint`",
        "  select_list <list>   a tappable list of up to 10 options",
        "  select_button <list> up to 3 tappable buttons",
        "  template             an approved template, and waits for a reply",
        "  note                 a message that expects no reply",
        "  begin group / end group    a branch; put the condition in `relevance`",
        "",
        "  `text` and `integer` look identical to a respondent. They differ "
        "entirely in the data: `text` stores whatever arrives, including "
        "'about 5'. Use `integer` whenever the answer is a number.",
        "",
        "  Prefer select_list. A tap cannot be malformed, so there is nothing to "
        "clean afterwards and nothing to guess about intent. Typing produces "
        "'3', '3.', 'tres', 'la 3', '3 por favor' - every one a real answer from "
        "a cooperative person, and every one needing a pattern written in "
        "advance or a hand-clean later.",
        "",
        "name - the variable name. Becomes the warehouse column.",
        "  Two rows may share a name if they are in different groups. That is "
        "how the same question asked two ways stays one question.",
        "",
        "label:<lang> - the message the respondent receives",
        "  One column per language. Alt+Enter for a line break. Aim at about 250 "
        "characters and 7 lines: that is the median across IPA's WhatsApp "
        "instruments, and a question longer than its screen pushes its own "
        "options out of view - which shows up as drift in the answers, never as "
        "an error.",
        "",
        "role - leave BLANK for an ordinary question",
        "  Only the spine positions need one: "
        + ", ".join(sorted(r for r in ROLES if r)),
        "  `consent` is a question like any other and is written like one, with "
        "its wording in `label` and its options in `choices`. It has a role only "
        "because it re-asks once rather than twice, and because an unreadable "
        "reply is recorded as unclear rather than as a refusal.",
        "",
        "relevance - when this row applies, e.g. ${arm}='2'",
        "  On a `begin group` row it branches the instrument. On a `close` row it "
        "picks which closing an outcome gets.",
        "",
        "retries - how many times to re-ask a reply that does not validate",
        f"  Blank uses {spec.settings.default_retries}. Only meaningful where a "
        "reply can fail, so a select or a constrained number - `text` cannot "
        "fail. Re-asking somebody who cannot answer is badgering a volunteer, so "
        "this is an ethical setting as much as a technical one.",
        "",
        "constraint - the regex a reply must match, for integer and decimal",
        "  Blank uses the default for the type. Whatever you write here is RUN "
        "against real replies when you check the spec, not merely read.",
        "",
        "constraint_message - which nudge to send when a reply does not validate",
        "  A key from the messages sheet. Blank picks a sensible one.",
        "",
        f"timeout - seconds to wait for a reply. Blank uses "
        f"{spec.settings.default_timeout}.",
        "",
        "stop_check - default yes, and it should stay yes",
        "  Routes a mid-survey STOP before the reply is treated as an answer. "
        "Twilio's own opt-out handling covers the carrier keywords for SMS; "
        "inside a WhatsApp session a STOP is an ordinary message and nothing "
        "looks at it unless the flow does. Without this, saying stop stores "
        "'stop' as the answer and the next question goes out anyway.",
        "",
        "publish - default yes. Whether the answer reaches the warehouse.",
        "encrypt - default no. Encrypt this field before it is stored.",
        "  For direct identifiers only - a name, a phone number.",
    ]


def _help_choices(spec: Spec) -> list[str]:
    """Return the guidance for the choices sheet."""
    return [
        "The choices worksheet",
        "",
        "One row per answer option. `list_name` ties a block of rows to the "
        "question that uses them: `select_list p1` reads the rows whose "
        "list_name is p1, in this order.",
        "",
        "value - the CODE stored in the warehouse, not the text",
        "  Set it explicitly on every row. A 'Prefer not to say' left to its "
        "position would code as a 6 on a 5-point scale and be averaged in "
        "silently - give it -99, or whatever your codebook uses.",
        "",
        "option_id - what a tapped row actually sends back",
        "  Leave it blank and one is generated. Worth knowing why it exists: a "
        "tapped LIST ROW returns this id, while a tapped BUTTON returns its "
        "visible title. The two interactive types disagree, and a flow that "
        "expected the label sent every list tap to the retry nudge - a real "
        "round lost an entire arm to it.",
        "",
        "label:<lang> - what the respondent sees. 24 CHARACTERS.",
        "  Short because WhatsApp shows no more, and this is the limit that "
        "shapes question design rather than just implementation. Every standard "
        "Likert label fits except the neutral midpoint: 'Neither agree nor "
        "disagree' is 26. Reword it.",
        "",
        "  NO EMOJI in a label. It is compared literally against the reply, and "
        "a variation selector makes two identical-looking strings different "
        "strings. Warmth goes in the question body, which nothing matches on.",
        "",
        "  Commas are fine. They were not always: the conditions used to take "
        "their alternatives as one comma-delimited string, so a comma split a "
        "label into two alternatives that never matched.",
        "",
        "description:<lang> - a second line under the label. 72 characters.",
        "",
        "typed:<lang> - extra spellings to accept, separated by |",
        "  Only wired up for the consent list at the moment. Anywhere else it "
        "would be accepted here and refused by the flow, so `check` says so.",
        "",
        "A list picker takes 1 to 10 options. Buttons take 3, because these are "
        "never submitted to Meta and WhatsApp permits 3 in session - past that "
        "the send fails rather than truncating.",
        "",
        "Note 0-10 is eleven points, so an NPS item does not fit a list picker.",
    ]


def _help_messages(spec: Spec) -> list[str]:
    """Return the guidance for the messages sheet."""
    return [
        "The messages worksheet",
        "",
        "Strings with no position in the instrument - a nudge sent when a reply "
        "cannot be read, the words that mean stop. Anything that DOES have a "
        "position is a row in the survey sheet instead, including every closing "
        "message: those are `note` rows whose `relevance` picks the outcome they "
        "belong to.",
        "",
        "Keys this instrument uses: " + ", ".join(m.key for m in spec.messages),
        "",
        "error_option / error_option_labels - the nudge after an unreadable "
        "reply to a select. Two of them because a question whose options are "
        "themselves numbers must NOT invite 'reply with the number': on a "
        "0 / 1 / 2-3 projects scale a typed '1' could be the position or the "
        "label, they are different options, and the flow refuses the digit "
        "rather than guess. Inviting it would ask for precisely the reply that "
        "cannot be accepted. The right one is chosen automatically.",
        "",
        "error_numeric - the nudge after a reply that failed a constraint.",
        "",
        "stop_words - comma-separated. Include the English words even in a "
        "non-English instrument: people type STOP whatever language they are "
        "answering in.",
        "",
        "list_button - the up-to-20-character button that opens a list.",
        "",
        "unsolicited - sent to somebody who writes to the number without having "
        "been launched into the survey. In practice that is mostly people being "
        "polite after finishing, and re-sending the opener to them would restart "
        "the whole survey.",
        "",
        "{button} in a nudge is filled in with the list button's text, so a "
        "nudge cannot name a button that is not on screen.",
    ]


def _write_help(workbook: Workbook, spec: Spec) -> None:
    """Write the instruction sheets, the way SurveyCTO ships its own."""
    for title, lines in (
        (f"help-{SURVEY_SHEET}", _help_survey(spec)),
        (f"help-{CHOICES_SHEET}", _help_choices(spec)),
        (f"help-{MESSAGES_SHEET}", _help_messages(spec)),
    ):
        sheet = workbook.create_sheet(title)
        sheet.column_dimensions["A"].width = 100
        for line in lines:
            sheet.append([line])
        for index, cell in enumerate(sheet["A"], start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if index == 1:
                cell.font = Font(bold=True, size=14)
            elif cell.value and not str(cell.value).startswith(" "):
                # A heading is a line that introduces a column, so it ends
                # without a full stop and starts hard against the margin.
                cell.font = Font(bold=not str(cell.value).endswith("."))


# ---------------------------------------------------------------------------
# Reading - the direction that matters, because this is the editing surface.
# ---------------------------------------------------------------------------


def read_xlsx(path: Path) -> Spec:
    """Read a spec back out of a workbook.

    Args:
        path: The ``.xlsx`` to read.

    Returns:
        The spec it describes.

    Raises:
        SpecError: If the file cannot be read or has no ``survey`` sheet.

    """
    path = Path(path)
    try:
        # dtype=str throughout, matching every other reader in this package.
        # Without it `1-2 times` comes back as a Timestamp and `007` as 7.
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    except Exception as exc:
        raise SpecError(f"Could not read {path}: {exc}") from exc

    if SURVEY_SHEET not in sheets:
        raise SpecError(
            f"{path} has no {SURVEY_SHEET!r} sheet, so it is not a survey "
            f"workbook. Found: {', '.join(sheets) or 'nothing'}"
        )

    settings = _read_settings(sheets.get(SETTINGS_SHEET))
    survey_frame = sheets[SURVEY_SHEET]
    languages = _languages(list(survey_frame.columns), "label") or settings.languages
    if not settings.languages:
        settings.languages = languages

    return Spec(
        settings=settings,
        survey=_read_survey(survey_frame, settings.languages),
        choices=_read_choices(sheets.get(CHOICES_SHEET), settings.languages),
        messages=_read_messages(sheets.get(MESSAGES_SHEET), settings.languages),
    )


def _read_settings(frame: pd.DataFrame | None) -> Settings:
    """Read the key/value settings sheet."""
    settings = Settings()
    if frame is None or frame.empty:
        return settings

    values = {
        _clean_cell(row.get("key")): _clean_cell(row.get("value"))
        for _, row in frame.iterrows()
    }

    settings.form_id = values.get("form_id", "")
    settings.form_title = values.get("form_title", "")
    settings.languages = [
        part.strip() for part in values.get("languages", "").split(",") if part.strip()
    ]
    settings.default_language = values.get("default_language", "") or (
        settings.languages[0] if settings.languages else "en"
    )
    settings.functions_service = values.get("functions_service", "") or "rtt-survey"
    settings.default_timeout = _parse_int(values.get("default_timeout", "")) or 3600
    retries = _parse_int(values.get("default_retries", ""))
    settings.default_retries = 2 if retries is None else retries
    settings.preloads = [
        part.strip() for part in values.get("preloads", "").split(",") if part.strip()
    ]

    for key, value in values.items():
        stem, _, lang = key.partition(":")
        if not lang or not value:
            continue
        if stem == "flow_name":
            settings.flow_name[lang] = value
        elif stem == "description":
            settings.description[lang] = value

    return settings


def _read_survey(frame: pd.DataFrame, languages: list[str]) -> list[SurveyRow]:
    """Read the survey sheet, skipping rows an RA left entirely blank."""
    rows: list[SurveyRow] = []
    for _, raw in frame.iterrows():
        row = {key: raw.get(key) for key in frame.columns}
        row_type = _clean_cell(row.get("type"))
        name = _clean_cell(row.get("name"))
        # A wholly empty row is Excel's, not the author's - a spreadsheet grows
        # trailing blanks as soon as anybody scrolls it.
        if not row_type and not name:
            continue

        rows.append(
            SurveyRow(
                type=row_type,
                name=name,
                role=_clean_cell(row.get("role")),
                relevance=_clean_cell(row.get("relevance")),
                label=_text_from(row, "label", languages),
                list_button=_text_from(row, "list_button", languages),
                template=_text_from(row, "template", languages),
                retries=_parse_int(row.get("retries")),
                timeout=_parse_int(row.get("timeout")),
                constraint=_clean_cell(row.get("constraint")),
                constraint_message=_clean_cell(row.get("constraint_message")),
                stop_check=_parse_bool(row.get("stop_check"), True),
                publish=_parse_bool(row.get("publish"), True),
                encrypt=_parse_bool(row.get("encrypt"), False),
            )
        )
    return rows


def _read_choices(frame: pd.DataFrame | None, languages: list[str]) -> list[ChoiceRow]:
    """Read the choices sheet."""
    if frame is None or frame.empty:
        return []
    rows: list[ChoiceRow] = []
    for _, raw in frame.iterrows():
        row = {key: raw.get(key) for key in frame.columns}
        list_name = _clean_cell(row.get("list_name"))
        value = _clean_cell(row.get("value"))
        if not list_name and not value:
            continue
        rows.append(
            ChoiceRow(
                list_name=list_name,
                value=value,
                option_id=_clean_cell(row.get("option_id")),
                label=_text_from(row, "label", languages),
                description=_text_from(row, "description", languages),
                typed=_text_from(row, "typed", languages),
            )
        )
    return rows


def _read_messages(
    frame: pd.DataFrame | None, languages: list[str]
) -> list[MessageRow]:
    """Read the messages sheet."""
    if frame is None or frame.empty:
        return []
    rows: list[MessageRow] = []
    for _, raw in frame.iterrows():
        row = {key: raw.get(key) for key in frame.columns}
        key_name = _clean_cell(row.get("key"))
        if not key_name:
            continue
        rows.append(MessageRow(key=key_name, text=_text_from(row, "text", languages)))
    return rows


#: Re-exported so a caller can say what it means without importing two modules.
__all__ = [
    "CONSTRAINED_TYPES",
    "read_xlsx",
    "write_xlsx",
]

"""Tests for the survey spec: the schema, its validator, and its two forms.

Most of these are about the *validator* rather than the schema, because the
schema is only as useful as its refusals. A spec that reads back cleanly and
then compiles into a flow where one option can never be matched is worse than no
spec at all: it moves the defect from a canvas somebody might squint at into a
spreadsheet everybody trusts.

The round-trip tests are the other half. The workbook is the editing surface, so
an edit that does not survive the trip back to JSON is silent data loss on the
instrument - and Excel is an actively hostile serialisation format, so each of
those cases is pinned individually rather than assumed to be covered by the
whole-spec comparison.
"""

import json
from pathlib import Path

import pytest

from requests_to_twilio.spec import (
    DEFAULT_CONSTRAINTS,
    ChoiceRow,
    MessageRow,
    Settings,
    Spec,
    SpecError,
    SurveyRow,
    check_spec,
    load_spec,
    review_notes,
    save_spec,
    spec_from_dict,
    spec_to_dict,
)
from requests_to_twilio.spec_xlsx import read_xlsx, write_xlsx

DEMO = Path(__file__).resolve().parents[1] / "surveys" / "data_use_demo.json"


def minimal_spec(**overrides) -> Spec:
    """A valid two-question instrument, for tests that break one thing at a time."""
    spec = Spec(
        settings=Settings(
            form_id="t",
            form_title="Test",
            languages=["en"],
            default_language="en",
        ),
        survey=[
            SurveyRow(
                type="select_button consent",
                name="consent",
                role="consent",
                label={"en": "Take part?"},
                retries=1,
            ),
            SurveyRow(
                type="select_list colours",
                name="Q1",
                label={"en": "Pick a colour"},
                retries=2,
            ),
        ],
        choices=[
            ChoiceRow(
                list_name="consent",
                value="yes",
                option_id="consent_yes",
                label={"en": "Yes"},
                typed={"en": "y"},
            ),
            ChoiceRow(
                list_name="consent",
                value="no",
                option_id="consent_no",
                label={"en": "No"},
                typed={"en": "n"},
            ),
            ChoiceRow(
                list_name="colours", value="1", option_id="c_red", label={"en": "Red"}
            ),
            ChoiceRow(
                list_name="colours", value="2", option_id="c_blue", label={"en": "Blue"}
            ),
        ],
        messages=[MessageRow(key="error_option", text={"en": "Try again"})],
    )
    for attribute, value in overrides.items():
        setattr(spec, attribute, value)
    return spec


class TestTheMinimalSpecIsActuallyValid:
    """If the fixture were invalid, every test below would pass for free."""

    def test_it_passes_its_own_checks(self):
        assert check_spec(minimal_spec()) == []


class TestJsonIsTheCanonicalForm:
    def test_a_spec_survives_a_json_round_trip(self):
        spec = minimal_spec()
        assert spec_to_dict(spec_from_dict(spec_to_dict(spec))) == spec_to_dict(spec)

    def test_defaults_are_not_written_out(self):
        """The tracked JSON should say what is unusual, not restate the obvious."""
        payload = spec_to_dict(minimal_spec())
        consent = payload["survey"][0]
        assert "stop_check" not in consent
        assert "publish" not in consent
        assert "encrypt" not in consent

    def test_a_non_default_flag_is_written_out(self):
        spec = minimal_spec()
        spec.survey[0].publish = False
        spec.survey[0].encrypt = True
        consent = spec_to_dict(spec)["survey"][0]
        assert consent["publish"] is False
        assert consent["encrypt"] is True

    def test_reading_a_missing_file_says_so(self, tmp_path):
        with pytest.raises(SpecError, match="Could not read"):
            load_spec(tmp_path / "nope.json")

    def test_reading_something_that_is_not_a_spec_says_so(self, tmp_path):
        path = tmp_path / "bad.json"
        path.write_text("[1, 2, 3]", encoding="utf-8")
        with pytest.raises(SpecError, match="must be a JSON object"):
            load_spec(path)

    def test_saving_creates_the_directory(self, tmp_path):
        path = tmp_path / "nested" / "deeper" / "spec.json"
        save_spec(minimal_spec(), path)
        assert (
            json.loads(path.read_text(encoding="utf-8"))["settings"]["form_id"] == "t"
        )


class TestOneRowIsOneSubgraph:
    """The claim the format makes. If the arithmetic is wrong, the claim is."""

    @pytest.mark.parametrize(
        ("row", "expected"),
        [
            # The opener stores nothing: any reply opens the 24-hour window.
            (SurveyRow(type="template", name="i", stop_check=False), 1),
            # ask, stop-check, store.
            (SurveyRow(type="text", name="q"), 3),
            # No stop check is one widget fewer, and a bad idea.
            (SurveyRow(type="text", name="q", stop_check=False), 2),
            # ask, stop-check, validate, store, count, decide, nudge, give up.
            (SurveyRow(type="select_list l", name="q", retries=2), 8),
            # A validated number is the same shape as a select.
            (SurveyRow(type="integer", name="q", retries=2), 8),
            # `numeric` is an alias, so it must count the same.
            (SurveyRow(type="numeric", name="q", retries=2), 8),
            # Zero retries drops the whole retry machinery.
            (SurveyRow(type="select_list l", name="q", retries=0), 4),
            (SurveyRow(type="note", name="c"), 1),
            # A group is only a widget when it branches.
            (SurveyRow(type="begin group", name="g", relevance="${arm}='1'"), 1),
            (SurveyRow(type="begin group", name="g"), 0),
        ],
    )
    def test_widget_count_per_row_type(self, row, expected):
        assert Spec().widget_count(row) == expected

    def test_text_cannot_be_re_asked_however_the_sheet_is_filled_in(self):
        """`text` accepts anything, so there is nothing for a retry to act on."""
        spec = Spec()
        assert spec.retries_for(SurveyRow(type="text", name="q", retries=3)) == 0

    def test_the_demo_matches_the_flow_it_came_from(self):
        """Counted against the real definition: intro 1, consent 7, arms 3 and 8."""
        spec = load_spec(DEMO)
        counts = {
            (row.name, row.type): spec.widget_count(row)
            for row in spec.survey
            if row.is_question
        }
        assert counts[("intro", "template")] == 1
        assert counts[("consent", "select_button consent")] == 7
        assert counts[("P1", "text")] == 3
        assert counts[("P1", "select_list p1")] == 8


class TestTheValidatorRefusesWhatItShould:
    def test_an_option_with_no_label_is_caught(self):
        """An unlabelled option is unreachable: there is nothing to tap or type.

        Note what this does *not* test. `_check_options_are_matchable` runs every
        label through the generated pattern, but both are now derived from the
        same choices sheet, so they cannot disagree by construction and no spec
        can currently trigger it. It is a regression guard for the day somebody
        makes the pattern configurable - which is exactly how the original defect
        happened, when the pattern expected a label and a tap sent an id.
        """
        spec = minimal_spec()
        spec.choices[2].label = {"en": ""}
        assert any("no label" in p for p in check_spec(spec))

    def test_a_numeric_scale_does_not_accept_an_ambiguous_bare_digit(self):
        """`1` on a 0 / 1-2 scale is the position or the label, and they differ.

        Passing cleanly is the correct outcome, and the interesting one:
        `positions_are_ambiguous` sees label 0 sitting at position 1, so
        `answer_pattern` drops the digit alternative entirely. A bare digit then
        matches nothing and the respondent is asked again - which is right for an
        input that genuinely has two readings.
        """
        from requests_to_twilio.answers import answer_pattern
        from requests_to_twilio.flows import evaluate_condition

        spec = minimal_spec()
        spec.choices[2].label = {"en": "0 times"}
        spec.choices[3].label = {"en": "1-2 times"}
        assert check_spec(spec) == []

        pattern = answer_pattern(spec.options("colours", "en"))
        assert not evaluate_condition("regex", pattern, "1")
        assert evaluate_condition("regex", pattern, "0 times")
        assert evaluate_condition("regex", pattern, "c_red")

    def test_an_over_long_option_label_is_caught(self):
        spec = minimal_spec()
        spec.choices[2].label = {"en": "Neither agree nor disagree"}  # 26 chars
        assert any("cap is 24" in p for p in check_spec(spec))

    def test_an_emoji_in_a_label_is_caught(self):
        """Compared literally against the reply, so a variation selector breaks it."""
        spec = minimal_spec()
        spec.choices[2].label = {"en": "Red 🔴"}
        assert any("emoji" in p for p in check_spec(spec))

    def test_an_emoji_in_a_body_is_fine(self):
        """Nothing matches on the body, so warmth belongs there."""
        spec = minimal_spec()
        spec.survey[1].label = {"en": "📊 Pick a colour"}
        assert check_spec(spec) == []

    def test_a_missing_choice_list_is_caught(self):
        spec = minimal_spec()
        spec.survey[1].type = "select_list nonexistent"
        assert any("no list called" in p for p in check_spec(spec))

    def test_a_select_without_a_list_name_is_caught(self):
        spec = minimal_spec()
        spec.survey[1].type = "select_list"
        assert any("needs a list name" in p for p in check_spec(spec))

    def test_an_unclosed_group_is_caught(self):
        spec = minimal_spec()
        spec.survey.insert(0, SurveyRow(type="begin group", name="G"))
        assert any("never closed" in p for p in check_spec(spec))

    def test_an_unopened_group_is_caught(self):
        spec = minimal_spec()
        spec.survey.append(SurveyRow(type="end group", name="G"))
        assert any("without a matching begin group" in p for p in check_spec(spec))

    def test_a_two_word_structural_type_is_not_reported_as_unknown(self):
        """`kind` takes the first word, which is right for select_list and wrong here."""
        spec = minimal_spec()
        spec.survey.insert(0, SurveyRow(type="begin group", name="G"))
        spec.survey.append(SurveyRow(type="end group", name="G"))
        assert not any("unknown type" in p for p in check_spec(spec))

    def test_the_same_name_in_two_groups_is_allowed(self):
        """ARM1_P1 and ARM2_P1 are the same question asked two ways."""
        spec = minimal_spec()
        spec.survey = [
            spec.survey[0],
            SurveyRow(type="begin group", name="A", relevance="${arm}='1'"),
            SurveyRow(type="text", name="P1", label={"en": "one way"}),
            SurveyRow(type="end group", name="A"),
            SurveyRow(type="begin group", name="B", relevance="${arm}='2'"),
            SurveyRow(type="text", name="P1", label={"en": "the other"}),
            SurveyRow(type="end group", name="B"),
        ]
        assert not any("already used" in p for p in check_spec(spec))

    def test_the_same_name_twice_in_one_group_is_caught(self):
        spec = minimal_spec()
        spec.survey.append(SurveyRow(type="text", name="Q1", label={"en": "again"}))
        assert any("already used" in p for p in check_spec(spec))

    def test_an_unknown_role_is_caught(self):
        spec = minimal_spec()
        spec.survey[1].role = "cosnent"
        assert any("unknown role" in p for p in check_spec(spec))

    def test_typed_outside_the_consent_list_is_caught(self):
        """The sheet would promise a spelling the generated pattern refuses."""
        spec = minimal_spec()
        spec.choices[2].typed = {"en": "rojo"}
        assert any("only wired up for the consent list" in p for p in check_spec(spec))

    def test_retries_on_a_text_question_is_caught(self):
        spec = minimal_spec()
        spec.survey.append(
            SurveyRow(type="text", name="Q2", label={"en": "anything"}, retries=2)
        )
        assert any("cannot fail" in p for p in check_spec(spec))

    def test_a_constraint_on_a_select_is_caught(self):
        spec = minimal_spec()
        spec.survey[1].constraint = r"(?:\s*\d+\s*)"
        assert any(
            "does not validate a reply against one" in p for p in check_spec(spec)
        )

    def test_a_constraint_message_that_names_nothing_is_caught(self):
        spec = minimal_spec()
        spec.survey[1].constraint_message = "no_such_message"
        assert any("not a key in the messages sheet" in p for p in check_spec(spec))

    def test_too_many_buttons_is_caught(self):
        """WhatsApp permits 3 in session, and past that the send fails."""
        spec = minimal_spec()
        for extra in range(4):
            spec.choices.append(
                ChoiceRow(
                    list_name="consent",
                    value=f"x{extra}",
                    label={"en": f"Option {extra}"},
                )
            )
        assert any("permits 3 buttons" in p for p in check_spec(spec))

    def test_more_than_ten_list_options_is_caught(self):
        spec = minimal_spec()
        for extra in range(3, 15):
            spec.choices.append(
                ChoiceRow(
                    list_name="colours", value=str(extra), label={"en": f"C{extra}"}
                )
            )
        assert any("list picker allows 1 to 10" in p for p in check_spec(spec))

    def test_a_duplicated_option_label_is_caught(self):
        """Two identical labels make the stored code ambiguous."""
        spec = minimal_spec()
        spec.choices[3].label = {"en": "Red"}
        assert any("repeats the label" in p for p in check_spec(spec))

    def test_a_missing_label_for_a_second_language_is_caught(self):
        spec = minimal_spec()
        spec.settings.languages = ["en", "es"]
        assert any("no label for language 'es'" in p for p in check_spec(spec))

    def test_an_instrument_with_no_consent_is_refused(self):
        spec = minimal_spec()
        spec.survey = [spec.survey[1]]
        assert any("role 'consent'" in p for p in check_spec(spec))

    def test_a_consent_reply_matching_both_branches_is_caught(self):
        """Otherwise participation is decided by transition order."""
        spec = minimal_spec()
        spec.choices[1].typed = {"en": "y"}
        assert any("matches both" in p for p in check_spec(spec))

    def test_a_template_row_needs_a_template_name(self):
        spec = minimal_spec()
        spec.survey.insert(0, SurveyRow(type="template", name="intro", role="intro"))
        assert any("names none for language" in p for p in check_spec(spec))

    def test_a_template_row_needs_no_label(self):
        """Its copy lives in templates/<name>.json, which is its only home."""
        spec = minimal_spec()
        spec.survey.insert(
            0,
            SurveyRow(
                type="template", name="intro", role="intro", template={"en": "t_intro"}
            ),
        )
        assert check_spec(spec) == []


class TestConstraintsAreRunNotRead:
    @pytest.mark.parametrize("reply", ["0", "3", "12", " 7 ", "100"])
    def test_the_integer_default_accepts_a_number(self, reply):
        spec = minimal_spec()
        spec.survey.append(
            SurveyRow(type="integer", name="N", label={"en": "How many?"}, retries=2)
        )
        assert check_spec(spec) == []

    def test_the_integer_default_is_unsigned(self):
        """A negative count would move a published mean, silently."""
        from requests_to_twilio.flows import evaluate_condition

        pattern = DEFAULT_CONSTRAINTS["integer"]
        assert evaluate_condition("regex", pattern, "5")
        assert not evaluate_condition("regex", pattern, "-5")
        assert not evaluate_condition("regex", pattern, "about 5")
        assert not evaluate_condition("regex", pattern, "3.5")

    def test_the_decimal_default_takes_a_comma_or_a_point(self):
        """Which one a respondent types is their locale, not the instrument's."""
        from requests_to_twilio.flows import evaluate_condition

        pattern = DEFAULT_CONSTRAINTS["decimal"]
        assert evaluate_condition("regex", pattern, "3.5")
        assert evaluate_condition("regex", pattern, "3,5")
        assert not evaluate_condition("regex", pattern, "3.5.5")

    def test_a_constraint_that_rejects_a_valid_answer_is_caught(self):
        """Too strict strands every respondent who answered correctly."""
        spec = minimal_spec()
        spec.survey.append(
            SurveyRow(
                type="integer",
                name="N",
                label={"en": "How many?"},
                # Single digits only, so "12" is refused.
                constraint=r"(?:\s*\d\s*)",
                retries=2,
            )
        )
        assert any("rejects '12'" in p for p in check_spec(spec))

    def test_a_custom_constraint_is_not_second_guessed_on_what_it_allows(self):
        """Somebody writing their own pattern may well mean to allow a negative."""
        spec = minimal_spec()
        spec.survey.append(
            SurveyRow(
                type="integer",
                name="N",
                label={"en": "Change in income"},
                constraint=r"(?:\s*-?\d+\s*)",
                retries=2,
            )
        )
        assert check_spec(spec) == []

    def test_numeric_is_an_alias_for_integer(self):
        spec = minimal_spec()
        spec.survey.append(
            SurveyRow(type="numeric", name="N", label={"en": "How many?"}, retries=2)
        )
        assert check_spec(spec) == []
        assert spec.survey[-1].resolved_constraint() == DEFAULT_CONSTRAINTS["integer"]


class TestConsentIsFlaggedForHumanReview:
    """The one string in the instrument that no check can judge."""

    def test_the_wording_is_surfaced(self):
        notes = review_notes(minimal_spec())
        assert len(notes) == 1
        assert "Take part?" in notes[0]
        assert "by hand" in notes[0]

    def test_review_notes_are_not_failures(self):
        """They must not gate a build, or they become the thing to switch off."""
        spec = minimal_spec()
        assert review_notes(spec)
        assert check_spec(spec) == []

    def test_an_instrument_without_consent_has_nothing_to_review(self):
        assert review_notes(Spec()) == []


class TestTheWorkbookRoundTrip:
    """The workbook is the editing surface, so this direction is load-bearing."""

    def test_a_spec_survives_the_trip_through_excel(self, tmp_path):
        spec = minimal_spec()
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert spec_to_dict(read_xlsx(tmp_path / "s.xlsx")) == spec_to_dict(spec)

    def test_the_demo_survives_the_trip_through_excel(self, tmp_path):
        """Two languages, emoji, commas in labels, numeric option labels."""
        spec = load_spec(DEMO)
        write_xlsx(spec, tmp_path / "demo.xlsx")
        assert spec_to_dict(read_xlsx(tmp_path / "demo.xlsx")) == spec_to_dict(spec)

    def test_the_round_tripped_demo_still_passes_its_checks(self, tmp_path):
        spec = load_spec(DEMO)
        write_xlsx(spec, tmp_path / "demo.xlsx")
        assert check_spec(read_xlsx(tmp_path / "demo.xlsx")) == []

    def test_the_help_sheets_ship_inside_the_file(self, tmp_path):
        import pandas as pd

        write_xlsx(minimal_spec(), tmp_path / "s.xlsx")
        sheets = pd.read_excel(tmp_path / "s.xlsx", sheet_name=None, dtype=str)
        assert {"survey", "choices", "messages", "settings"} <= set(sheets)
        assert {"help-survey", "help-choices", "help-messages"} <= set(sheets)

    def test_reading_something_that_is_not_a_survey_workbook_says_so(self, tmp_path):
        import pandas as pd

        path = tmp_path / "notasurvey.xlsx"
        pd.DataFrame({"a": ["b"]}).to_excel(path, index=False)
        with pytest.raises(SpecError, match="no 'survey' sheet"):
            read_xlsx(path)


class TestExcelSpecificHazards:
    """Each of these silently changes the instrument, and none is visible."""

    def test_a_label_that_looks_like_a_date_stays_text(self, tmp_path):
        """`1-2 times` is coerced to a date in several locales."""
        spec = minimal_spec()
        spec.choices[2].label = {"en": "1-2 times"}
        spec.choices[3].label = {"en": "3-5 times"}
        write_xlsx(spec, tmp_path / "s.xlsx")
        back = read_xlsx(tmp_path / "s.xlsx")
        assert [c.label["en"] for c in back.choice_list("colours")] == [
            "1-2 times",
            "3-5 times",
        ]

    def test_a_negative_code_stays_a_string(self, tmp_path):
        """-99 for 'Prefer not to say' must not come back as a number."""
        spec = minimal_spec()
        spec.choices.append(
            ChoiceRow(
                list_name="colours",
                value="-99",
                option_id="c_refuse",
                label={"en": "Prefer not to say"},
            )
        )
        write_xlsx(spec, tmp_path / "s.xlsx")
        back = read_xlsx(tmp_path / "s.xlsx")
        assert back.choice_list("colours")[-1].value == "-99"

    def test_a_leading_zero_code_keeps_its_zero(self, tmp_path):
        """The same lesson `push_file` learned: an identifier is not a number."""
        spec = minimal_spec()
        spec.choices[2].value = "007"
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert read_xlsx(tmp_path / "s.xlsx").choice_list("colours")[0].value == "007"

    def test_newlines_in_a_body_survive(self, tmp_path):
        spec = minimal_spec()
        spec.survey[1].label = {"en": "Question 1 of 4\n\nPick a colour"}
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert (
            read_xlsx(tmp_path / "s.xlsx").row("Q1").label["en"]
            == "Question 1 of 4\n\nPick a colour"
        )

    def test_a_windows_newline_is_normalised(self, tmp_path):
        """An edit in Excel on Windows can hand back \\r\\n, changing the message."""
        spec = minimal_spec()
        spec.survey[1].label = {"en": "Line one\r\n\r\nLine two"}
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert "\r" not in read_xlsx(tmp_path / "s.xlsx").row("Q1").label["en"]

    def test_trailing_whitespace_is_stripped(self, tmp_path):
        """Invisible in a cell, load-bearing in a literally-compared label."""
        spec = minimal_spec()
        spec.choices[2].label = {"en": "Red  "}
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert (
            read_xlsx(tmp_path / "s.xlsx").choice_list("colours")[0].label["en"]
            == "Red"
        )

    def test_emoji_in_a_body_survive(self, tmp_path):
        spec = minimal_spec()
        spec.survey[1].label = {"en": "📊 Pick a colour 🙌"}
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert read_xlsx(tmp_path / "s.xlsx").row("Q1").label["en"] == (
            "📊 Pick a colour 🙌"
        )

    def test_a_comma_in_a_label_survives(self, tmp_path):
        """It broke the conditions once; it must not break the workbook now."""
        spec = minimal_spec()
        spec.choices[2].label = {"en": "Yes, always"}
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert (
            read_xlsx(tmp_path / "s.xlsx").choice_list("colours")[0].label["en"]
            == "Yes, always"
        )

    def test_a_yes_no_flag_round_trips(self, tmp_path):
        spec = minimal_spec()
        spec.survey[1].stop_check = False
        spec.survey[1].publish = False
        spec.survey[1].encrypt = True
        write_xlsx(spec, tmp_path / "s.xlsx")
        back = read_xlsx(tmp_path / "s.xlsx").row("Q1")
        assert (back.stop_check, back.publish, back.encrypt) == (False, False, True)

    def test_a_retry_count_of_zero_is_not_read_as_blank(self, tmp_path):
        """Zero and unset mean different things: none, versus take the default."""
        spec = minimal_spec()
        spec.survey[1].retries = 0
        write_xlsx(spec, tmp_path / "s.xlsx")
        assert read_xlsx(tmp_path / "s.xlsx").row("Q1").retries == 0

    def test_blank_trailing_rows_are_ignored(self, tmp_path):
        """A spreadsheet grows empty rows as soon as anyone scrolls it."""
        import openpyxl

        write_xlsx(minimal_spec(), tmp_path / "s.xlsx")
        book = openpyxl.load_workbook(tmp_path / "s.xlsx")
        book["survey"].append([None] * 8)
        book["survey"].append([None] * 8)
        book.save(tmp_path / "s.xlsx")
        assert len(read_xlsx(tmp_path / "s.xlsx").survey) == 2

    def test_flags_are_blank_where_they_mean_nothing(self, tmp_path):
        """A `stop_check: yes` on a group row is noise, not information."""
        import openpyxl

        spec = load_spec(DEMO)
        write_xlsx(spec, tmp_path / "demo.xlsx")
        sheet = openpyxl.load_workbook(tmp_path / "demo.xlsx")["survey"]
        head = [c.value for c in sheet[1]]

        by_name = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            by_name[(row[head.index("type")], row[head.index("name")])] = row

        # openpyxl reads a genuinely empty cell as None rather than "".
        group = by_name[("begin group", "ARM1")]
        for flag in ("stop_check", "publish", "encrypt"):
            assert group[head.index(flag)] in (None, ""), (
                f"{flag} should be blank on a group row"
            )

        # But a non-default value stays, because there it is the point: a closing
        # message publishes no column, and that is worth seeing.
        closing = by_name[("note", "close_complete")]
        assert closing[head.index("publish")] == "no"
        assert closing[head.index("stop_check")] in (None, "")

        # And a real question still says all three.
        question = by_name[("select_list p1", "P1")]
        assert question[head.index("stop_check")] == "yes"
        assert question[head.index("publish")] == "yes"
        assert question[head.index("encrypt")] == "no"

    def test_blanking_those_flags_does_not_break_the_round_trip(self, tmp_path):
        """Blank reads back as the default, which is what was blanked."""
        spec = load_spec(DEMO)
        write_xlsx(spec, tmp_path / "demo.xlsx")
        assert spec_to_dict(read_xlsx(tmp_path / "demo.xlsx")) == spec_to_dict(spec)
